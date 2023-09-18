import openpyxl
import requests
from bs4 import BeautifulSoup
import os
import time



# Path to your URLs Excel file
url_excel_path = "C:\\Users\\alont\\OneDrive\\ezway\\links.xlsx"

def download_and_get_soup(url):
    response = requests.get(url)
    if response.status_code == 200:
        return BeautifulSoup(response.content, 'html.parser')
    else:
        print(f'Failed to download the file. HTTP Status Code: {response.status_code}')
        return None

def extract_company_details(soup):
    company_name_tag = soup.find("span", {"id": "HeaderEntityNameEB"})
    company_name = company_name_tag.text if company_name_tag else "N/A"
    
    company_number_tag = soup.find("span", {"id": "HeaderSingNumberD"})
    company_number = company_number_tag.text if company_number_tag else "N/A"
    
    return company_name, company_number

def extract_data_to_excel(soup, worksheet, row_idx, url):
    fields_map = {
        "Field212": "סה\"כ הון",
        "Field214": "סה\"כ הון והתחייבויות",
        "Field225": "הכנסות",
        "Field229": "רווח גולמי",
        "Field284": "רווח (הפסד) לפני מס",
        "Field294": "רווח (הפסד)",
        "Field297": "רווח (הפסד) שניתן לייחוס לבעלים של החברה האם",
        "Field299": "רווח (הפסד) שניתן לייחוס לזכויות שאינן מקנות שליטה",
        "Field460": "סה\"כ רווח (הפסד) בסיסי למניה",
        "Field462": "סה\"כ רווח (הפסד) מדולל למניה",
        "Field632": "רווח כולל שניתן לייחוס לבעלים של החברה האם",
        "Field634": "רווח כולל שניתן לייחוס לזכויות שאינן מקנות שליטה",
        "Field408": "תזרימי מזומנים, נטו שנבעו (ששימשו) מפעילויות שוטפות",
        "Field410": "תזרימי מזומנים, נטו שנבעו (ששימשו) מפעילויות השקעה",
        "Field412": "תזרימי מזומנים,נטו שנבעו (ששימשו) מפעילויות מימון",
        "Field414": "השפעת שינויים בשער חליפין של מטבע חוץ על מזומנים ושווי מזומנים",
        "Field659": "השפעות נוספות שלא קיבלו ביטוי בסעיפים 4-1 לעיל",
        "Field648": "עלייה (ירידה), נטו במזומנים ושווי מזומנים בתקופה",
        "Field650": "יתרת מזומנים ושווי מזומנים לתחילת תקופה",
        "Field652": "יתרת מזומנים ושווי מזומנים לסוף תקופה",
        "Field653": "סוג הדוח",
        "Field25": "תקופה"
    }

    for col_idx, (field_id, field_name) in enumerate(fields_map.items(), start=2):
        value_tag = soup.find("span", {"id": field_id})
        value = value_tag.text if value_tag else 'N/A'
        worksheet.cell(row=row_idx, column=col_idx).value = value

    # Extracting company details
    company_name, company_number = extract_company_details(soup)
    worksheet.cell(row=row_idx, column=col_idx + 1).value = company_name
    worksheet.cell(row=row_idx, column=col_idx + 2).value = company_number

# Read URLs from the Excel file
url_workbook = openpyxl.load_workbook(url_excel_path)
url_worksheet = url_workbook.active
urls = [row[1].value for row in url_worksheet.iter_rows(min_row=2) if row[1].value]

# Initialize new Excel workbook for the metrics
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active

# Writing headers to the worksheet
headers_extended = ["Website URL", "סה\"כ הון", "סה\"כ הון והתחייבויות", "הכנסות", "סוג הדוח", "תקופה", "שם החברה", "מספר ברשם"]
for col_idx, header in enumerate(headers_extended, 1):
    output_worksheet.cell(row=1, column=col_idx).value = header

# Extract data for each URL
for row_idx, url in enumerate(urls, start=2):
    print(f'Processing {url}...')
    soup = download_and_get_soup(url)
    if soup:
        output_worksheet.cell(row=row_idx, column=1).value = url
        extract_data_to_excel(soup, output_worksheet, row_idx, url)
    time.sleep(0.1)


# Save the compiled data
compiled_file_path = "C:\\Users\\alont\\OneDrive\\פרויקט עם אלון\\compiled_metrics35 חברות.xlsx"
output_workbook.save(compiled_file_path)
print(f"Saved data to {compiled_file_path}")



