import pyodbc
import schedule
import time
from datetime import datetime
import openpyxl


# Load DB details from keys.txt
with open('C:/Users/alont/OneDrive/פרויקט עם אלון/keys[2].txt', 'r') as file:
    lines = file.readlines()
    db_details = {line.split('=')[0]: line.split('=')[1].strip() for line in lines}

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active

# Use the STOCKHISTORY function in cell B1 to get historical data from A1 (stock symbol)
ws['B1'] = '=STOCKHISTORY(A1, TODAY(), TODAY())'

wb.save('stock_history.xlsx')

def update_stock_price(stock_name):
    # Open the existing Excel file
    wb = openpyxl.load_workbook('stock_history.xlsx')
    ws = wb.active

    # Insert the stock symbol into Excel
    ws['A1'] = f'"{stock_name}"'
    wb.save('stock_history.xlsx')
    
    # Assume Excel calculates the price automatically in cell B1
    # Here we simulate it for the example
    stock_price = ws['B1'].value
    
    # Connect to the database and update the stock price
    conn_str = (
        f"Driver={{ODBC Driver 17 for SQL Server}};"
        f"Server={db_details['server']};"
        f"Database={db_details['database']};"
        f"UID={db_details['username']};"
        f"PWD={db_details['password']};"
    )
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    update_query = f"UPDATE stocks_prices SET StockPrice = ?, LastUpdated = ? WHERE Stock_name = ?"
    last_updated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor.execute(update_query, (stock_price, last_updated, stock_name))
    conn.commit()

    cursor.close()
    conn.close()
    wb.close()

def check_and_update_stock_prices():
    # Connect to the database and read stock names
    conn_str = (
        f"Driver={{ODBC Driver 17 for SQL Server}};"
        f"Server={db_details['server']};"
        f"Database={db_details['database']};"
        f"UID={db_details['username']};"
        f"PWD={db_details['password']};"
    )
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    cursor.execute("SELECT Stock_name FROM stocks_prices")
    
    stock_names = [row.Stock_name for row in cursor.fetchall()]

    for stock_name in stock_names:
        update_stock_price(stock_name)

    conn.close()

# Schedule the function to run every 30 minutes
schedule.every(30).minutes.do(check_and_update_stock_prices)

while True:
    schedule.run_pending()
    time.sleep(1)
