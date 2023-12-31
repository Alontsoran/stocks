   
import pyodbc
import schedule
import time
from datetime import datetime
<<<<<<< HEAD
import openpyxl
import os

# Load DB details from keys.txt
with open('C:/Users/zvi25/Desktop/StartUp/פרויקט עם אלון/Milestones/2/2.1/keys.txt', 'r', encoding='utf-8') as file:
=======
import certifi
import json
import ssl
import concurrent.futures

try:
    from urllib.request import urlopen
except ImportError:
    from urllib2 import urlopen

def get_stock_price(stock_name):
    print(f"Getting stock price for {stock_name}")
    url = f"https://financialmodelingprep.com/api/v3/quote-short/{stock_name}?apikey=728c1b4c2d8218add598fcfc7401f70b"
    
    context = ssl.create_default_context(cafile=certifi.where())
    
    try:
        response = urlopen(url, context=context)
        data = response.read().decode("utf-8")
        stock_data = json.loads(data)
        
        if stock_data:
            latest_close_price = float(stock_data[0]['price'])
            return latest_close_price
    except Exception as e:
        print(f"Failed to get stock price for {stock_name} due to {type(e).__name__}: {e}")
        return None

with open('C:/Users/alont/OneDrive/פרויקט עם אלון/keys.txt', 'r') as file:
>>>>>>> 6270f501c0e6ae5f63af1cff4cc02c6c6e30c397
    lines = file.readlines()
    db_details = {line.split('=')[0]: line.split('=')[1].strip() for line in lines}

conn_str = (
    f"Driver={{ODBC Driver 17 for SQL Server}};"
    f"Server={db_details['server']};"
    f"Database={db_details['database']};"
    f"UID={db_details['username']};"
    f"PWD={db_details['password']};"
)
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

<<<<<<< HEAD
# Use the STOCKHISTORY function in cell B1 to get historical data from A1 (stock symbol)
ws['B1'] = '=STOCKHISTORY(A1, TODAY(),TODAY(),0,0,1)'

desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
file_path = os.path.join(desktop_path, 'stock_history.xlsx')
wb.save(file_path)

def update_stock_price(stock_name):
    # Open the existing Excel file
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Insert the stock symbol into Excel
    ws['A1'] = f'"{stock_name}"'
    wb.save(file_path)
    
    # Assume Excel calculates the price automatically in cell B2 (or whatever cell the value is in)
    stock_price = ws['B2'].value  # Change this to the appropriate cell if it's not B2
    
    # Print the current value from Excel for verification
    print(f"Current value for {stock_name}: {stock_price}")

    # Ensure the value is float
    stock_price = float(stock_price)
    
    # Connect to the database and update the stock price
    conn_str = (
        f"Driver={{ODBC Driver 17 for SQL Server}};"
        f"Server={db_details['server']};"
        f"Database={db_details['database']};"
        f"UID={db_details['username']};"
        f"PWD={db_details['password']};"
    )
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    
    # Get the current date and time in the format "dd_mm_yyyy_hh_mm"
    current_datetime = datetime.now().strftime("%d_%m_%Y_%H_%M")
    
    # Add a new column without checking if it exists
    add_column_query = f"ALTER TABLE stocks_prices ADD [{current_datetime}] FLOAT"
    cursor.execute(add_column_query)

    # Update the stock price for the current datetime
    update_query = f"UPDATE stocks_prices SET [{current_datetime}] = ? WHERE Stock_name = ?"
    cursor.execute(update_query, (stock_price, stock_name))
    conn.commit()
=======
def update_stock_prices():
    global cursor, conn
    print("Updating stock prices")
    current_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    column_name = f"price_{current_time}"
    
    alter_table_query = f"ALTER TABLE stockp ADD [{column_name}] FLOAT"
    print(f"Executing alter_table_query at {current_time}")
    cursor.execute(alter_table_query)
    
    cursor.execute("SELECT names FROM stockp")
    stock_names = [row.names for row in cursor.fetchall()]
    
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = []
        for stock_name in stock_names:
            stock_name_with_suffix = stock_name.strip() + ".TA"
            futures.append(executor.submit(get_stock_price, stock_name_with_suffix))
        
        for future, stock_name in zip(concurrent.futures.as_completed(futures), stock_names):
            try:
                stock_price = future.result()
                print(f"Stock price for {stock_name} is {stock_price}")
                if stock_price is not None:
                    update_query = f"UPDATE stockp SET [{column_name}] = ? WHERE names = ?"
                    print(f"Executing update_query for {stock_name} at {current_time}")
                    cursor.execute(update_query, (stock_price, stock_name))
                    conn.commit()
                    print(f"Database has been updated with the latest price for {stock_name}: {stock_price}")
            except Exception as e:
                print(f"Failed to update stock price for {stock_name}: {e}")

print("Script has started. Waiting for the first scheduled run...")
schedule.every(1).minutes.do(update_stock_prices)
>>>>>>> 6270f501c0e6ae5f63af1cff4cc02c6c6e30c397

try:
    while True:
        schedule.run_pending()
        time.sleep(1)
except KeyboardInterrupt:
    print("Stopping the scheduler...")
finally:
    cursor.close()
    conn.close()
<<<<<<< HEAD
    wb.close()

def check_and_update_stock_prices():
    # Connect to the database and read stock names
    conn_str = (
        f"Driver={{ODBC Driver 17 for SQL Server}};"
        f"Server={db_details['server']};"
        f"Database={db_details['database']};"
        f"UID={db_details['username']};"
        f"PWD={db_details['password']};"
    )
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    cursor.execute("SELECT Stock_name FROM stocks_prices")
    
    stock_names = [row.Stock_name for row in cursor.fetchall()]

    for stock_name in stock_names:
        update_stock_price(stock_name)

    conn.close()

# Schedule the function to run every 30 minutes
schedule.every(0.1).minutes.do(check_and_update_stock_prices)

while True:
    schedule.run_pending()
    time.sleep(1)
=======
>>>>>>> 6270f501c0e6ae5f63af1cff4cc02c6c6e30c397
